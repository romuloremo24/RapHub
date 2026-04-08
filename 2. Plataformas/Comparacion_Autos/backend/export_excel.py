"""Excel exporter – generates a formatted .xlsx after each scrape run."""
from __future__ import annotations
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

EXPORTS_DIR = Path(__file__).parent / "exports"

# Light background per source for row tinting
_SOURCE_COLORS = {
    "mercadolibre": "FFFDE7",   # yellow-50
    "chileautos":   "FFEBEE",   # red-50
    "kavak":        "E8F5E9",   # green-50
    "yapo":         "FFF3E0",   # orange-50
    "clicar":       "FCE4EC",   # pink-50
    "facebook":     "E3F2FD",   # blue-50
    "autocosmos":   "EDE7F6",   # purple-50
    "autosusados":  "F3E5F5",   # deep-purple-50
    "autocl":       "E0F7FA",   # cyan-50
    "gildemeister": "E1F5FE",   # light-blue-50
    "economicos":   "EFEBE9",   # brown-50
    "autojusto":    "F1F8E9",   # light-green-50
}
_DEFAULT_COLOR = "F5F5F5"
_MISSING_COLOR = "FFCDD2"   # red-100 for empty required fields
_HEADER_COLOR  = "1D52CC"   # primary blue

# Columns: (header_label, dict_key, width, is_required)
_COLUMNS = [
    ("Fecha scrape",   "scraped_date",  14, False),
    ("Plataforma",     "source",        14, False),
    ("Título",         "title",         40, True),
    ("Marca",          "brand",         14, True),
    ("Modelo",         "model",         16, True),
    ("Año",            "year",          8,  True),
    ("KM",             "km",            10, True),
    ("Precio",         "price",         14, True),
    ("Moneda",         "currency",      8,  False),
    ("Condición",      "condition",     10, False),
    ("Combustible",    "fuel_type",     13, False),
    ("Transmisión",    "transmission",  13, False),
    ("Ubicación",      "location",      18, False),
    ("Tiene foto",     "_has_photo",    10, False),
    ("URL",            "url",           45, False),
    ("URL foto",       "image_url",     45, False),
]


def export_scrape_to_excel(run_date: str, listings: list, summary: dict) -> Optional[Path]:
    """
    Create exports/scrape_YYYY-MM-DD.xlsx with:
      - Sheet "Listings"  – one row per listing, colour-coded by source, missing fields in red
      - Sheet "Resumen"   – per-source completeness stats
    Returns the path to the created file, or None on error.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed – run: pip install openpyxl")
        return None

    if not listings:
        logger.info("No listings to export for %s", run_date)
        return None

    EXPORTS_DIR.mkdir(exist_ok=True)
    out_path = EXPORTS_DIR / f"scrape_{run_date}.xlsx"

    wb = Workbook()

    # ── Sheet 1: Listings ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Listings"
    ws.freeze_panes = "A2"

    # Header row
    header_fill = PatternFill("solid", fgColor=_HEADER_COLOR)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (label, _, width, _req) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, listing in enumerate(listings, start=2):
        source = listing.get("source", "")
        bg = PatternFill("solid", fgColor=_SOURCE_COLORS.get(source, _DEFAULT_COLOR))
        missing_fill = PatternFill("solid", fgColor=_MISSING_COLOR)

        for col_idx, (_, key, _, is_required) in enumerate(_COLUMNS, start=1):
            if key == "_has_photo":
                raw = listing.get("image_url")
                value = "✓" if raw else "✗"
            else:
                raw = listing.get(key)
                value = raw

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

            # Highlight missing required fields
            if is_required and (raw is None or raw == "" or raw == 0):
                cell.fill = missing_fill
            else:
                cell.fill = bg

            # Format price as number
            if key == "price" and isinstance(value, (int, float)):
                cell.number_format = '#,##0'

            # Format KM as number
            if key == "km" and isinstance(value, (int, float)):
                cell.number_format = '#,##0'

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}1"

    # ── Sheet 2: Resumen ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.freeze_panes = "A2"

    res_headers = [
        "Plataforma", "Total", "Sin marca", "Sin modelo", "Sin año",
        "Sin km", "Sin precio", "Sin foto", "Con error", "Duración (s)",
    ]
    for col_idx, h in enumerate(res_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = 15

    # Compute per-source stats from listings
    stats: dict = {}
    for l in listings:
        src = l.get("source", "unknown")
        if src not in stats:
            stats[src] = {"total": 0, "no_brand": 0, "no_model": 0, "no_year": 0,
                          "no_km": 0, "no_price": 0, "no_photo": 0}
        s = stats[src]
        s["total"] += 1
        if not l.get("brand"):     s["no_brand"] += 1
        if not l.get("model"):     s["no_model"] += 1
        if not l.get("year"):      s["no_year"]  += 1
        if not l.get("km"):        s["no_km"]    += 1
        if not l.get("price"):     s["no_price"] += 1
        if not l.get("image_url"): s["no_photo"] += 1

    pct_fill = PatternFill("solid", fgColor="FFF9C4")  # yellow for high-miss %
    for row_idx, src in enumerate(sorted(stats), start=2):
        s = stats[src]
        n = s["total"] or 1
        src_summary = summary.get("sources", {}).get(src, {})
        values = [
            src, s["total"],
            f"{s['no_brand']}/{n} ({100*s['no_brand']//n}%)",
            f"{s['no_model']}/{n} ({100*s['no_model']//n}%)",
            f"{s['no_year']}/{n} ({100*s['no_year']//n}%)",
            f"{s['no_km']}/{n} ({100*s['no_km']//n}%)",
            f"{s['no_price']}/{n} ({100*s['no_price']//n}%)",
            f"{s['no_photo']}/{n} ({100*s['no_photo']//n}%)",
            src_summary.get("error") or "–",
            src_summary.get("duration", "–"),
        ]
        bg = PatternFill("solid", fgColor=_SOURCE_COLORS.get(src, _DEFAULT_COLOR))
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            # Highlight high missing %
            if isinstance(val, str) and "%" in val:
                try:
                    pct = int(val.split("(")[1].replace("%)", ""))
                    cell.fill = pct_fill if pct > 30 else bg
                except Exception:
                    cell.fill = bg
            else:
                cell.fill = bg

    # ── Sheet 3: Log ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Log")
    ws3.cell(row=1, column=1, value="Generado").font = Font(bold=True)
    ws3.cell(row=1, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws3.cell(row=2, column=1, value="Fecha scrape").font = Font(bold=True)
    ws3.cell(row=2, column=2, value=run_date)
    ws3.cell(row=3, column=1, value="Total listings").font = Font(bold=True)
    ws3.cell(row=3, column=2, value=len(listings))
    ws3.cell(row=4, column=1, value="Nuevos (DB)").font = Font(bold=True)
    ws3.cell(row=4, column=2, value=summary.get("total_new", "–"))
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 25

    wb.save(out_path)
    logger.info("Excel exported: %s (%d listings)", out_path.name, len(listings))
    return out_path
